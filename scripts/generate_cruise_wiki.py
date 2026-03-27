#!/usr/bin/env python3
"""
Cruise Process Wiki Generator — Royal Caribbean Group reference
Calls Qwen via Ollama → sanitises Mermaid → renders PNG via mmdc
→ builds HTML pages with sidebar + swimlanes + L4 table
→ pushes to GitHub (ghatk047/Cruise-Process-Wiki)

Usage:
  python3 generate_cruise_wiki.py                        # full batch
  python3 generate_cruise_wiki.py --pilot                # first 2 queued only
  python3 generate_cruise_wiki.py --max 10               # limit count
  python3 generate_cruise_wiki.py --start GS-EM-01       # resume from ID
  python3 generate_cruise_wiki.py --only GS-EM-01        # single process
  python3 generate_cruise_wiki.py --model qwen2.5:latest # faster model
  python3 generate_cruise_wiki.py --reset GS-EM-01,GS-EM-02
  python3 generate_cruise_wiki.py --reset-all
"""

import argparse, base64, json, os, re, subprocess, sys, time
from datetime import datetime
from html import escape
from pathlib import Path

try:
    import requests
except ImportError:
    os.system("pip3 install requests --break-system-packages -q"); import requests

try:
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    os.system("pip3 install openpyxl --break-system-packages -q")
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────
OLLAMA_URL  = "http://localhost:11434/api/generate"
GH_TOKEN    = os.environ.get("GH_TOKEN", "YOUR_TOKEN_HERE")  # set via env or replace
GH_REPO     = "ghatk047/Cruise-Process-Wiki"
GH_API      = "https://api.github.com"
BASE_DIR    = Path.home() / "Projects/cruise-wiki"
EXCEL_PATH  = BASE_DIR / "data/Cruise_Process_Catalog.xlsx"
JSON_PATH   = BASE_DIR / "data/processes.json"
DIAG_DIR    = BASE_DIR / "diagrams"
IMG_DIR     = BASE_DIR / "assets/img"

SWIM_COLOURS = [
    "#3b82f6","#8b5cf6","#ec4899","#f59e0b",
    "#10b981","#06b6d4","#6366f1","#0ea5e9",
    "#84cc16","#ef4444","#f97316","#14b8a6"
]

# ─────────────────────────────────────────────────────────────────────────────
# TAXONOMY  (11 L1 × 2-3 L2 × 8-10 L3 = 220 processes)
# Each tuple: (L1 name, L1 code, L1 slug, [
#   (L2 name, L2 code, L2 slug, [L3 process names...])
# ])
# ─────────────────────────────────────────────────────────────────────────────
TAXONOMY = [

  ("Guest Services & Embarkation", "GS", "guest-services", [
    ("Embarkation & Check-In", "GS-EM", "embarkation-checkin", [
      "Pre-Cruise Documentation & eDocs Processing",
      "Terminal Check-In & SeaPass Card Issuance",
      "Luggage Tagging & Drop-Off Management",
      "Muster 2.0 Safety Drill Completion & Verification",
      "Suite & Loyalty Member Priority Boarding",
      "Group Embarkation Coordination",
      "Accessibility & Special Needs Embarkation",
      "Minors & Unaccompanied Youth Check-In",
      "Pre-Cruise Upsell & Package Activation at Pier",
      "Embarkation Day Capacity & Flow Management",
    ]),
    ("Debarkation & Port Operations", "GS-DB", "debarkation", [
      "Debarkation Planning & Luggage Tag Distribution",
      "Express Self-Assist Debarkation Management",
      "Customs & Border Protection (CBP) Coordination",
      "Facial Recognition Debarkation (IDEMIA MFACE)",
      "Onward Travel & Transfer Coordination",
      "Last-Minute Overboard Guest Management",
      "Group Debarkation Coordination",
      "Lost Luggage & Left-Behind Property Handling",
      "Port Agent Briefing & Services Coordination",
      "In-Transit Port Stop Guest Management",
    ]),
    ("Guest Experience & Concierge", "GS-CX", "guest-experience", [
      "Guest Services Desk Operations",
      "Royal App AI Chat & Self-Service",
      "Guest Complaint Resolution & Service Recovery",
      "Accessibility & PRM Guest Support",
      "Medical Centre Triage & Guest Health Services",
      "Onboard Account Query & Billing Resolution",
      "Lost & Found Management",
      "Crown & Anchor Loyalty Recognition Onboard",
      "VIP & Suite Concierge Royal Genie Service",
      "Inter-Cruise Booking & Next Voyage Sales",
    ]),
  ]),

  ("Stateroom & Housekeeping", "HK", "housekeeping", [
    ("Stateroom Operations", "HK-SR", "stateroom-ops", [
      "Daily Stateroom Cleaning & Turndown Service",
      "Suite & Royal Suite Class Premium Service",
      "Stateroom Inspection & Quality Scoring",
      "Linen & Towel Management & Par Stock",
      "Out-of-Service Cabin Tracking & Maintenance Routing",
      "VIP & Special Occasion Amenity Placement",
      "Stateroom Preventive Maintenance Coordination",
      "Embarkation Day Deep Clean & Turnaround",
    ]),
    ("Laundry & Public Areas", "HK-LP", "laundry-public", [
      "Shipboard Laundry Operations & Processing",
      "Guest Laundry & Pressing Service",
      "Uniform Laundering & Crew Issue Management",
      "Public Area & Corridor Cleaning Schedule",
      "Pool Deck & Lido Area Maintenance",
      "Gym & Fitness Area Cleaning & Upkeep",
      "Kids Club & Youth Zone Cleaning",
      "Deep Cleaning & Periodic Sanitation Programme",
    ]),
  ]),

  ("Food & Beverage Operations", "FB", "food-beverage", [
    ("Restaurant & Dining Management", "FB-RD", "restaurant-dining", [
      "Main Dining Room (MDR) Service Operations",
      "Specialty Restaurant Operations & Reservations",
      "Windjammer Buffet Setup & Operations",
      "Room Service Delivery & Order Management",
      "Beverage Package Activation & Bar Operations",
      "Kosher & Special Dietary Accommodation",
      "Celebrity Chef & Signature Restaurant Management",
      "Restaurant Revenue & Cover Count Reporting",
    ]),
    ("Culinary & Provisioning", "FB-CP", "culinary-provisioning", [
      "Menu Engineering & Recipe Costing",
      "Food Provisioning Order & Vendor Management",
      "Cold Chain & HACCP Compliance at Sea",
      "Galley Production Planning & Execution",
      "Food Waste Monitoring & Reduction (50% goal)",
      "AI-Driven Inventory Optimisation (Symphony platform)",
      "Allergen Management & Dietary Labelling",
      "Inspection & Quality Audit of F&B Operations",
    ]),
  ]),

  ("Entertainment & Activities", "EN", "entertainment", [
    ("Showroom & Live Entertainment", "EN-SH", "showroom-entertainment", [
      "Production Show Scheduling & Technical Setup",
      "Broadway & Headliner Show Operations",
      "AquaTheater Show Production (Oasis Class)",
      "Ice Show Production & Rink Management",
      "Comedy, Magic & Variety Act Management",
      "Guest Entertainer Onboarding & Contract Management",
      "Show Ticket Reservation & Capacity Management",
      "Post-Show Changeover & Technical Reset",
    ]),
    ("Activities & Adventure", "EN-AC", "activities-adventure", [
      "FlowRider Surf Simulator Operations",
      "Rock Climbing Wall & Zip Line Management",
      "Sports Court & Fitness Programme Scheduling",
      "Kids Club (Adventure Ocean) Daily Programme",
      "Teen Club & Youth Activity Management",
      "Casino Royale Operations & Club Royale Loyalty",
      "Bingo, Trivia & Onboard Games Programme",
      "Photography & Media Sales Operations",
    ]),
  ]),

  ("Shore Excursions & Destinations", "SE", "shore-excursions", [
    ("Shore Excursion Operations", "SE-OP", "excursion-operations", [
      "Shore Excursion Catalogue Management & Pricing",
      "Royal App & Web Pre-Booking Management",
      "Onboard Shore Excursion Desk Operations",
      "Local Tour Operator Vetting & Contracting",
      "Tour Bus & Transportation Coordination",
      "Tender Operations (Anchor-Port Management)",
      "Shore Excursion Cancellation & Refund Processing",
      "Post-Excursion Guest Feedback & Incident Reporting",
    ]),
    ("Private Destinations & CocoCay", "SE-PD", "private-destinations", [
      "Perfect Day at CocoCay Daily Operations",
      "Royal Beach Club Paradise Island Operations",
      "Water Park & Thrill Waterway Management",
      "Private Island F&B & Retail Operations",
      "Transportation from Ship to Island",
      "Private Cabana & Premium Experience Booking",
      "Environmental Compliance at Private Destinations",
      "Security & Guest Safety at Private Islands",
    ]),
  ]),

  ("Marine & Technical Operations", "MT", "marine-technical", [
    ("Bridge & Navigation", "MT-BN", "bridge-navigation", [
      "Voyage Planning & Route Optimisation",
      "Bridge Watch Keeping & OOW Procedures",
      "Port Approach, Docking & Undocking",
      "Pilotage Management & Port Authority Compliance",
      "Weather Routing & Storm Avoidance",
      "GMDSS & Emergency Communications Management",
      "Electronic Chart Display (ECDIS) Operations",
      "Vessel Traffic Service (VTS) Communication",
      "Passage Plan Review & Master's Approval",
      "Night Order & Standing Order Management",
    ]),
    ("Engineering & Propulsion", "MT-EP", "engineering-propulsion", [
      "Main Engine & Propulsion Plant Operations",
      "LNG Fuel Management (Icon Class Ships)",
      "Electrical Power Generation & Distribution",
      "Hotel Load & HVAC Systems Management",
      "Planned Maintenance System (PMS) Scheduling",
      "Machinery Space Rounds & Watch Management",
      "Fuel Consumption Monitoring & Efficiency Reporting",
      "Emergency Diesel Generator Testing & Readiness",
      "Bilge & Ballast System Management",
      "Refrigeration & Cold Store Plant Management",
    ]),
    ("Deck & Safety Management", "MT-DS", "deck-safety", [
      "Lifeboat & Life-Saving Appliance (LSA) Inspection",
      "Fire Detection System Testing & Maintenance",
      "ISPS Code Compliance & Ship Security Plan",
      "Man Overboard (MOB) Drill & Response",
      "Damage Control & Flooding Emergency Response",
      "Anchor & Mooring Operations Management",
      "Crane & Deck Equipment Maintenance",
      "Port State Control (PSC) Inspection Preparation",
      "Enclosed Space Entry Permit Management",
      "Dry Dock Planning & Shipyard Coordination",
    ]),
  ]),

  ("Environmental & Sustainability", "ES", "environmental", [
    ("Environmental Compliance", "ES-EC", "environmental-compliance", [
      "MARPOL Annex I — Oily Water Separator Management",
      "MARPOL Annex IV — Sewage Treatment Operations (AWTS)",
      "MARPOL Annex V — Garbage Management Plan Execution",
      "MARPOL Annex VI — Exhaust Gas Scrubber & Emissions",
      "Ballast Water Treatment System (BWTS) Operations",
      "Air Quality Monitoring in ECAs (0.1% Sulphur Compliance)",
      "Environmental Incident Reporting & Spill Response",
      "Annual Environmental Audit & Regulatory Submission",
    ]),
    ("Sustainability Programme", "ES-SP", "sustainability-programme", [
      "Food Waste Reduction Programme (Target: -50% by 2025)",
      "Single-Use Plastic Elimination Tracking",
      "Shore Power Connection (Cold Ironing) Operations",
      "Carbon Footprint Tracking & IMO CII Reporting",
      "Sustainable Procurement & Supplier Assessment",
      "Green Certification (EMAS, ISO 14001) Management",
      "Crew Environmental Awareness Training",
      "Destination Community & Coral Reef Protection",
    ]),
  ]),

  ("Crew Management & HR", "CM", "crew-management", [
    ("Crew Operations", "CM-CO", "crew-operations", [
      "Crew Joining & Sign-On Documentation",
      "Crew Immigration & Flag State Compliance",
      "Crew Medical Examination & Fitness-to-Work Clearance",
      "Crew Watch Schedule & Rotation Management",
      "Crew Welfare, Recreation & Internet Access",
      "Payroll & Allotment Processing (Flag State Rules)",
      "Crew Disciplinary & Grievance Management",
      "Crew Sign-Off, Repatriation & Travel Coordination",
    ]),
    ("Training & Certification", "CM-TC", "training-certification", [
      "STCW Basic Safety Training Verification",
      "Ship-Specific Familiarisation & Safety Induction",
      "Onboard Training Drill Programme (Fire, MOB, Flooding)",
      "HELM (Hospitality) Training Delivery",
      "Officer GMDSS Licence & Certificate Renewal",
      "Mandatory Medical First Aid Training",
      "Environmental Officer Training",
      "Performance Appraisal & Career Development Planning",
    ]),
  ]),

  ("Revenue & Commercial", "RC", "revenue-commercial", [
    ("Revenue Management", "RC-RM", "revenue-management", [
      "Cruise Fare Pricing & Demand Forecasting",
      "Dynamic Pricing & Last-Minute Offer Management",
      "Royal Up (Upgrade Bidding) Programme Management",
      "Beverage Package & Dining Package Yield Management",
      "Shore Excursion Revenue Optimisation",
      "Casino Revenue & Hold Percentage Management",
      "Onboard Revenue Forecasting & Flash Reporting",
      "Revenue Performance Review vs Budget (Oracle OBIEE)",
    ]),
    ("Sales, Reservations & Trade", "RC-SR", "sales-reservations", [
      "Direct Booking via Royalcaribbean.com & App (RES system)",
      "Travel Agent GDS Booking & Commission Management",
      "Group Booking & Charter Management",
      "CruiseTour & Land Package Booking",
      "Reservation Modification, Cancel & Refund Processing",
      "Pre-Cruise Upsell Campaign Management",
      "Corporate & Incentive Travel Programme Sales",
      "Trade Partner Portal & Marketing Co-Op Management",
    ]),
  ]),

  ("Finance & Procurement", "FN", "finance-procurement", [
    ("Onboard Finance & Accounting", "FN-OF", "onboard-finance", [
      "SeaPass Onboard Account Management & Billing",
      "Daily Folio Audit & Revenue Reconciliation",
      "Casino Cage Operations & Count Room Procedures",
      "Foreign Currency Exchange & Cash Management",
      "End-of-Voyage Financial Close & Revenue Reporting",
      "Accounts Receivable — Travel Agent Commission Settlement",
      "Tax Compliance — Port-by-Port Sales Tax Management",
      "Passenger Compensation & Goodwill Credit Issuance",
    ]),
    ("Procurement & Supply Chain", "FN-PS", "procurement-supply", [
      "Ship Provisioning Order & Consolidated Purchasing",
      "Vendor Qualification & Preferred Supplier Management",
      "Port Agent Selection & Services Contracting",
      "Fuel Bunkering Operations & Price Hedging",
      "Spare Parts & Technical Stores Management",
      "Customs & Duty-Free Goods Management",
      "SAP Ariba Purchase Order Workflow",
      "Supplier Performance Review & Savings Reporting",
    ]),
  ]),

  ("Technology & Cybersecurity", "IT", "technology", [
    ("Shipboard IT & Connectivity", "IT-SC", "shipboard-it", [
      "Starlink Satellite Connectivity Management (98% Fleet)",
      "Shipboard Network & Wi-Fi Hotspot Operations",
      "Oracle Hospitality Cruise SPMS Administration",
      "Point-of-Sale (POS) System Management (Micros)",
      "Royal App Backend Services & API Management",
      "Onboard CCTV & Computer Vision (Microsoft Partnership)",
      "IT Service Desk — Crew & Guest Support",
      "Voyage Data Recorder (VDR) Maintenance",
      "Digital Signage & Wayfinding System Management",
      "Cashless Payment & SeaPass Terminal Management",
    ]),
    ("Cybersecurity & Data", "IT-CD", "cybersecurity-data", [
      "OT/IT Network Segmentation & Security",
      "Shipboard Cyber Incident Response",
      "PCI DSS Compliance for Onboard Payments",
      "Guest Data Privacy (GDPR/CCPA) Management",
      "Satellite Data Link Security & Encryption",
      "Vulnerability Assessment & Penetration Testing",
      "Oracle OBIEE / BI Analytics Platform Management",
      "Data Transfer Shore-Ship Synchronisation",
      "IT Disaster Recovery & Business Continuity",
      "Access Control & Identity Management (Crew/Guest)",
    ]),
  ]),

  ("Health, Safety & Medical", "HS", "health-safety", [
    ("Onboard Medical Operations", "HS-MD", "medical-operations", [
      "Medical Centre Daily Operations & Triage",
      "Emergency Medical Response & Resuscitation",
      "Guest Medical Evacuation (MEDEVAC) Coordination",
      "Infectious Disease Outbreak Management (Norovirus Protocol)",
      "Crew Medical Fitness & Routine Health Surveillance",
      "Pharmacy Management & Controlled Drug Records",
      "Medical Billing & Travel Insurance Claim Processing",
      "Medical Equipment Maintenance & Calibration",
      "Port Health Authority Reporting & Clearance",
      "Mental Health & Wellbeing Support (Crew & Guest)",
    ]),
    ("Safety & Emergency Management", "HS-SE", "safety-emergency", [
      "Safety Management System (SMS) Documentation & Audit",
      "Fire Safety Drill Planning & Execution",
      "Muster Station Accountability & Assembly Management",
      "Abandon Ship & Lifeboat Launch Procedures",
      "Search & Rescue (SAR) Coordination with Coast Guard",
      "Occupational Health & Safety Incident Reporting",
      "Crowd Management & Emergency Evacuation",
      "Serious Casualty Investigation & Root Cause Analysis",
      "Chemical & Hazardous Material Safety Management",
      "Food Safety & Outbreak Investigation Management",
    ]),
  ]),
]

# ─────────────────────────────────────────────────────────────────────────────
# Build flat process list and lookup map
# ─────────────────────────────────────────────────────────────────────────────
ALL_PROCS = []
PROC_MAP  = {}
for l1_name, l1_code, l1_slug, groups in TAXONOMY:
    for l2_name, l2_code, l2_slug, proc_names in groups:
        for i, pname in enumerate(proc_names, 1):
            pid = f"{l2_code}-{i:02d}"
            proc = {
                "id":        pid,
                "l1_domain": l1_name,
                "l1_code":   l1_code,
                "l1_slug":   l1_slug,
                "l2_process": l2_name,
                "l2_code":   l2_code,
                "l2_slug":   l2_slug,
                "l3_name":   pname,
                "status":    "Queued",
                "l4_steps":  [],
            }
            ALL_PROCS.append(proc)
            PROC_MAP[pid] = proc

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def ts():
    return datetime.now().strftime("%H:%M:%S")

def log(msg, kind="info"):
    icons = {"info":"   ","ok":"✅ ","warn":"⚠️  ","err":"❌ ","head":"🚀 ","step":"→  ","push":"📤 "}
    print(f"[{ts()}] {icons.get(kind,'   ')} {msg}", flush=True)

# ─────────────────────────────────────────────────────────────────────────────
# PROGRESS — processes.json
# ─────────────────────────────────────────────────────────────────────────────
def load_progress():
    if JSON_PATH.exists():
        try:
            saved = json.loads(JSON_PATH.read_text(encoding="utf-8"))
            for sp in saved.get("processes", []):
                pid = sp.get("id")
                if pid in PROC_MAP:
                    PROC_MAP[pid]["status"]   = sp.get("status", "Queued")
                    PROC_MAP[pid]["l4_steps"] = sp.get("l4_steps", [])
            log(f"Loaded progress from {JSON_PATH}")
        except Exception as ex:
            log(f"Could not load processes.json: {ex}", "warn")
    else:
        log("No processes.json found — starting fresh")

def save_progress():
    JSON_PATH.parent.mkdir(parents=True, exist_ok=True)
    data = {"processes": [{"id": p["id"], "status": p["status"], "l4_steps": p.get("l4_steps",[])} for p in ALL_PROCS]}
    JSON_PATH.write_text(json.dumps(data, indent=2), encoding="utf-8")

# ─────────────────────────────────────────────────────────────────────────────
# QWEN PROMPT
# ─────────────────────────────────────────────────────────────────────────────
def build_prompt(proc):
    pid  = proc["id"]
    l1   = proc["l1_domain"]
    l2   = proc["l2_process"]
    l3   = proc["l3_name"]
    return f"""You are a senior cruise operations expert with detailed knowledge of
Royal Caribbean Group (RCCL) operations across its fleet of 69 ships globally.

Generate a complete structured process document for: {pid} — {l3}
Domain: {l1} > {l2}
Reference Ship: Icon of the Seas (largest passenger ship, 250,800 GT, 7,600 guests)
Reference Line: Royal Caribbean International (brand under Royal Caribbean Group)

Key Royal Caribbean systems:
RES (core reservation system — 9M+ lines RPG code), Oracle Hospitality Cruise SPMS
(Shipboard Property Management), Oracle MICROS (POS), Oracle OBIEE (BI/analytics),
Royal Caribbean App (AI chatbot NLP, 35% adoption increase), Starlink satellite
connectivity (98% fleet, sub-100ms latency), Microsoft Azure AI (computer vision
for crowd management), Salesforce CRM, SAP S/4HANA, SAP Ariba, Workday HCM,
IDEMIA MFACE (facial recognition debarkation), Crown and Anchor loyalty platform,
Club Royale casino loyalty platform, SeaPass onboard account system,
Perfect Day at CocoCay operations platform, Bravo Poker (poker management),
Environmental compliance systems (MARPOL, BWTS, AWTS), PMS (Planned Maintenance System)

Return ONLY valid JSON with NO markdown fences, NO extra text:
{{
  "description": "2-3 sentence operational description at Royal Caribbean",
  "trigger": "what initiates this process",
  "outcome": "what successful completion produces",
  "l4_steps": [
    {{
      "step": "1.1",
      "name": "Step name",
      "role": "Exact Royal Caribbean role e.g. Guest Services Officer, Executive Chef, Chief Engineer",
      "system": "Exact system e.g. Oracle SPMS, Royal App, SAP S/4HANA",
      "input": "Input document or data",
      "output": "Output document or deliverable",
      "kpi": "Metric with target e.g. check-in < 8 min per guest",
      "decision_point": "Y or N",
      "exception": "Y or N",
      "pain_point": "Real Royal Caribbean operational challenge at this step"
    }}
  ],
  "swim_lanes": [
    {{"role": "Role Title", "color": "#3b82f6", "steps": ["1.1","1.2"]}}
  ],
  "systems": ["Oracle SPMS","Royal App"],
  "kpis": ["Guest NPS > 72","Check-in time < 8 min"],
  "risks": ["Starlink connectivity loss during peak","Port agent delays"],
  "mermaid": "..."
}}

MERMAID RULES — ALL MUST BE FOLLOWED (violations cause mmdc render failure):
- Line 1 of mermaid field: %%{{init: {{'theme':'base','themeVariables':{{'fontSize':'12px'}}}}}}%%
- Line 2: flowchart LR
- NO YAML frontmatter above flowchart (never use ---config:--- syntax)
- Node IDs MUST start with a LETTER: NodeA, StepB, Init, Exec — NEVER 1.1 or 2.3
- Arrows MUST be --> (two dashes + greater-than) NEVER --gt or --&gt; or - ->
- Use 3 subgraphs: P1[Phase 1 - Initiation], P2[Phase 2 - Execution], P3[Phase 3 - Close]
- Node labels: plain text only, no & < > " characters inside brackets
- End with: style P1 fill:#dbeafe,stroke:#3b82f6
- End with: style P2 fill:#fef3c7,stroke:#f59e0b
- End with: style P3 fill:#d1fae5,stroke:#10b981

CONTENT RULES:
- l4_steps: exactly 10-12 steps in 3 phases (1.x, 2.x, 3.x)
- swim_lanes: list every unique role with steps they own
- Real Royal Caribbean roles: Guest Services Officer, Stateroom Attendant, F&B Manager, etc.
- systems: 4-6 real Royal Caribbean systems
- kpis: 4-6 measurable KPIs with targets
- risks: 3-5 RCCL-specific risks
- JSON only, absolutely no markdown, no explanation"""

# ─────────────────────────────────────────────────────────────────────────────
# CALL QWEN
# ─────────────────────────────────────────────────────────────────────────────
def call_qwen(prompt, model="qwen2.5-coder:14b", max_retries=3):
    for attempt in range(1, max_retries + 1):
        log(f"Calling {model} (attempt {attempt})...", "info")
        t0 = time.time()
        try:
            r = requests.post(
                OLLAMA_URL,
                json={"model": model, "prompt": prompt, "stream": False, "format": "json"},
                timeout=360
            )
            elapsed = int(time.time() - t0)
            if r.status_code != 200:
                log(f"HTTP {r.status_code}: {r.text[:200]}", "warn")
                time.sleep(3); continue
            raw = r.json().get("response", "").strip()
            log(f"Response in {elapsed}s ({len(raw)} chars)")
            # Strip optional markdown fences
            raw = re.sub(r'^```(?:json)?\s*', '', raw, flags=re.I)
            raw = re.sub(r'```\s*$', '', raw).strip()
            data = json.loads(raw)
            if "l4_steps" in data and len(data["l4_steps"]) >= 6:
                return data
            log("Response missing l4_steps or too short — retrying", "warn")
        except json.JSONDecodeError as ex:
            log(f"JSON parse error: {ex} — retrying", "warn")
        except Exception as ex:
            log(f"Attempt {attempt} failed: {ex}", "warn")
        if attempt < max_retries:
            time.sleep(3)
    return None

# ─────────────────────────────────────────────────────────────────────────────
# MERMAID SANITISER  — canonical version, handles all known Qwen output issues
# ─────────────────────────────────────────────────────────────────────────────
def _mmd_fallback():
    """A clean 3-phase fallback that always renders via mmdc."""
    return (
        "%%{init: {'theme':'base','themeVariables':{'fontSize':'12px'}}}%%\n"
        "flowchart LR\n"
        "  subgraph P1[Phase 1 - Initiation]\n"
        "    Init([Start]) --> Plan[Plan and Brief Crew]\n"
        "    Plan --> Auth{Authorised?}\n"
        "    Auth -->|Yes| Exec\n"
        "    Auth -->|No| Escalate[Escalate to Supervisor]\n"
        "    Escalate --> Auth\n"
        "  end\n"
        "  subgraph P2[Phase 2 - Execution]\n"
        "    Exec[Execute Process Steps] --> Check{Issues Found?}\n"
        "    Check -->|Yes| Resolve[Resolve and Log in Oracle SPMS]\n"
        "    Check -->|No| Doc[Document Results]\n"
        "    Resolve --> Doc\n"
        "  end\n"
        "  subgraph P3[Phase 3 - Close]\n"
        "    Doc --> Update[Update Systems and Notify]\n"
        "    Update --> Close([Process Complete])\n"
        "  end\n"
        "  style P1 fill:#dbeafe,stroke:#3b82f6\n"
        "  style P2 fill:#fef3c7,stroke:#f59e0b\n"
        "  style P3 fill:#d1fae5,stroke:#10b981"
    )

def sanitise_mmd(mmd):
    """Sanitise Mermaid text from Qwen before passing to mmdc."""
    if not mmd or len(mmd.strip()) < 15:
        return _mmd_fallback()

    # 1. Normalise line endings and strip code fences
    mmd = mmd.replace("\r\n", "\n").replace("\r", "\n").strip()
    mmd = re.sub(r'^```(?:mermaid)?\s*', '', mmd, flags=re.I)
    mmd = re.sub(r'```\s*$', '', mmd).strip()

    # 2. CRITICAL: Remove any YAML frontmatter (---config:--- syntax)
    #    This causes parse error on line 1 in mmdc CLI
    mmd = re.sub(r'^---\s*\n.*?^---\s*\n', '', mmd, flags=re.M | re.S)

    # 3. Fix broken arrow syntax Qwen produces (HTML entity leak)
    #    --gt  →  -->
    mmd = re.sub(r'--\s*gt\b', '-->', mmd)
    mmd = re.sub(r'--&gt;', '-->', mmd)
    mmd = re.sub(r'-\s*->\s*gt\b', '-->', mmd)
    mmd = re.sub(r'(==|\.--)>gt\b', r'\1>', mmd)

    # 4. Fix node IDs that start with a digit (invalid in Mermaid)
    #    e.g. 1.1[Label] → S1_1[Label]
    def fix_ids(line):
        return re.sub(
            r'(?<![A-Za-z_])(\d+\.\d+)(?=[\s\[\(\{]|-->|==>|-\.->|$)',
            lambda m: 'S' + m.group(1).replace('.', '_'),
            line
        )
    mmd = '\n'.join(fix_ids(line) for line in mmd.splitlines())

    # 5. Clean special characters ONLY inside node label brackets [ ] ( ) { }
    def clean_label(m):
        bopen, content, bclose = m.group(1), m.group(2), m.group(3)
        content = content.replace('&amp;', ' and ').replace('&', ' and ')
        content = content.replace('&lt;', 'lt').replace('&gt;', 'gt')
        content = content.replace('<', 'lt').replace('>', 'gt')
        content = content.replace('"', "'")
        content = re.sub(r'  +', ' ', content).strip()
        return bopen + content + bclose

    mmd = re.sub(r'(\[)([^\]]*?)(\])', clean_label, mmd)
    mmd = re.sub(r'(\()([^\)]*?)(\))', clean_label, mmd)
    mmd = re.sub(r'(\{)([^\}]*?)(\})', clean_label, mmd)

    # 6. Ensure %%{init}%% header is present
    if '%%{init:' not in mmd and '%%{ init:' not in mmd:
        mmd = "%%{init: {'theme':'base','themeVariables':{'fontSize':'12px'}}}%%\n" + mmd

    # 7. Remove blank lines between %%{init}%% and flowchart directive
    lines = mmd.splitlines()
    cleaned = []
    i = 0
    while i < len(lines):
        cleaned.append(lines[i])
        if '%%{init:' in lines[i] or '%%{ init:' in lines[i]:
            # Skip any blank lines immediately following
            j = i + 1
            while j < len(lines) and not lines[j].strip():
                j += 1
            i = j
        else:
            i += 1
    mmd = '\n'.join(cleaned)

    # 8. Ensure diagram type header
    if not re.search(r'^(flowchart|graph\s)', mmd, re.M | re.I):
        first_non_init = next(
            (l for l in mmd.splitlines() if l.strip() and '%%' not in l), None
        )
        if first_non_init:
            mmd = mmd.replace(first_non_init, "flowchart LR\n" + first_non_init, 1)
        else:
            mmd += "\nflowchart LR\n"

    # 9. Final sanity check — broken arrows still present → fallback
    if re.search(r'--gt|--&gt;', mmd):
        log("Mermaid still has broken arrows after sanitise — using fallback", "warn")
        return _mmd_fallback()

    return mmd

# ─────────────────────────────────────────────────────────────────────────────
# PNG GENERATION via mmdc
# ─────────────────────────────────────────────────────────────────────────────
def generate_png(pid, mmd_text):
    DIAG_DIR.mkdir(parents=True, exist_ok=True)
    IMG_DIR.mkdir(parents=True, exist_ok=True)

    mmd_path = DIAG_DIR / f"{pid.lower()}.mmd"
    png_path  = IMG_DIR  / f"{pid.lower()}.png"
    mmd_path.write_text(mmd_text, encoding="utf-8")

    cmd = ["mmdc", "-i", str(mmd_path), "-o", str(png_path),
           "-w", "1920", "-H", "1080", "--scale", "2", "--backgroundColor", "white"]
    log(f"Running mmdc for {pid}...", "info")
    try:
        r = subprocess.run(cmd, capture_output=True, text=True, timeout=90)
        if r.returncode != 0:
            log(f"mmdc failed: {r.stderr[:300]}", "warn")
            # Try fallback diagram
            mmd_path.write_text(_mmd_fallback(), encoding="utf-8")
            r2 = subprocess.run(cmd, capture_output=True, text=True, timeout=90)
            if r2.returncode != 0:
                log(f"mmdc fallback also failed — no PNG for {pid}", "err")
                return False
    except subprocess.TimeoutExpired:
        log(f"mmdc timed out for {pid}", "err"); return False
    except FileNotFoundError:
        log("mmdc not found — run: npm install -g @mermaid-js/mermaid-cli", "err"); return False

    if not png_path.exists() or png_path.stat().st_size < 1000:
        log(f"PNG too small or missing for {pid}", "err"); return False

    log(f"PNG generated: {png_path.stat().st_size/1024:.0f} KB", "ok")

    # Push PNG to GitHub
    png_bytes = png_path.read_bytes()
    encoded   = base64.b64encode(png_bytes).decode("ascii")
    gh_path   = f"assets/img/{pid.lower()}.png"
    url       = f"{GH_API}/repos/{GH_REPO}/contents/{gh_path}"
    headers   = {"Authorization": f"token {GH_TOKEN}", "Accept": "application/vnd.github.v3+json"}
    sha = None
    try:
        rg = requests.get(url, headers=headers, timeout=15)
        if rg.status_code == 200: sha = rg.json()["sha"]
    except: pass
    body = {"message": f"Add {pid} BPMN diagram PNG", "content": encoded}
    if sha: body["sha"] = sha
    for attempt in range(3):
        try:
            rp = requests.put(url, headers=headers, json=body, timeout=60)
            if rp.status_code in (200, 201):
                log(f"PNG pushed: {gh_path}", "ok"); return True
            log(f"PNG push {rp.status_code}", "warn")
        except Exception as ex:
            log(f"PNG push error: {ex}", "warn")
        time.sleep(2)
    log(f"PNG push failed for {pid}", "err"); return False

# ─────────────────────────────────────────────────────────────────────────────
# SYSTEM TAG HELPER
# ─────────────────────────────────────────────────────────────────────────────
def stag(system):
    s = system.lower()
    if "spms" in s or "oracle hospitality" in s: cls = "spms"
    elif "oracle" in s:          cls = "oracle"
    elif "royal app" in s:       cls = "royal-app"
    elif "res" == s.strip() or "res system" in s: cls = "res-sys"
    elif "micros" in s:          cls = "micros"
    elif "seapass" in s:         cls = "seapass"
    elif "sap ariba" in s:       cls = "ariba"
    elif "sap" in s:             cls = "sap"
    elif "workday" in s:         cls = "workday"
    elif "salesforce" in s:      cls = "salesforce"
    elif "starlink" in s:        cls = "starlink"
    elif "obiee" in s or "bi" in s.split(): cls = "obiee"
    elif "mermaid" in s:         cls = "custom"
    elif "crown" in s or "anchor" in s: cls = "loyalty"
    elif "club royale" in s:     cls = "club-royale"
    else:                        cls = "custom"
    return f'<span class="stag {cls}">{escape(system)}</span>'

def flag(val):
    v = str(val).strip().upper()
    return f'<span class="flag {"flag-y" if v == "Y" else "flag-n"}">{v}</span>'

# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR BUILDER
# NOTE: Built once per page from PROC_MAP state at time of HTML generation.
# The sidebar is STATIC HTML embedded in each page — it is NOT updated after
# each push (this was a major performance issue in airport wiki: 23 index
# pages pushed after every single process = 28 seconds per process).
# Instead, the sidebar shows status dots based on processes.json at render time.
# The home page and L1/L2 index pages ARE updated after each process.
# ─────────────────────────────────────────────────────────────────────────────
def build_sidebar(active_pid, depth=3):
    """
    depth=3 → process page (needs ../../.. prefix for links)
    depth=2 → L2 index page
    depth=1 → L1 index page
    depth=0 → home page
    """
    prefix = "../" * depth

    lines = []
    for l1_name, l1_code, l1_slug, groups in TAXONOMY:
        active_in_l1 = any(
            f"{l2_code}-{i+1:02d}" == active_pid
            for _, l2_code, _, procs in groups
            for i in range(len(procs))
        )
        oc = " open" if active_in_l1 else ""
        lines.append(f'  <div class="sidebar-section">')
        lines.append(f'    <div class="sidebar-domain{oc}" data-label="{escape(l1_name)}"><span class="domain-label">{escape(l1_name)}</span><span class="chevron">▶</span></div>')
        lines.append(f'    <div class="sidebar-l2{oc}">')
        for l2_name, l2_code, l2_slug, procs in groups:
            lines.append(f'      <a class="sidebar-l2-link" href="{prefix}{l1_slug}/{l2_slug}/">{escape(l2_name)}</a>')
            lines.append(f'      <div class="sidebar-l3">')
            for i, pname in enumerate(procs, 1):
                pid_   = f"{l2_code}-{i:02d}"
                saved  = PROC_MAP.get(pid_, {})
                dot    = "status-done" if saved.get("status") == "Complete" else "status-queue"
                ac     = " active" if pid_ == active_pid else ""
                lines.append(
                    f'        <a class="sidebar-l3-link{ac}" href="{prefix}{l1_slug}/{l2_slug}/{pid_.lower()}/">'
                    f'<span class="status-dot {dot}"></span>'
                    f'<span class="pid">{pid_}</span>{escape(pname)}</a>'
                )
            lines.append(f'      </div>')
        lines.append(f'    </div>')
        lines.append(f'  </div>')
    return "\n".join(lines)

# ─────────────────────────────────────────────────────────────────────────────
# SWIMLANE SECTION
# ─────────────────────────────────────────────────────────────────────────────
def build_swimlane_section(swim_lanes, l4_steps):
    if not swim_lanes:
        return ""
    colour_map = {lane["role"]: lane.get("color", SWIM_COLOURS[i % len(SWIM_COLOURS)])
                  for i, lane in enumerate(swim_lanes)}
    rows = []
    for i, lane in enumerate(swim_lanes):
        role  = lane["role"]
        color = lane.get("color", SWIM_COLOURS[i % len(SWIM_COLOURS)])
        steps = lane.get("steps", [])
        step_names = []
        for sn in steps:
            match = next((s["name"] for s in l4_steps if s.get("step") == sn), sn)
            step_names.append(f'<span class="swim-step-badge">{escape(sn)}: {escape(match)}</span>')
        steps_html = " ".join(step_names) if step_names else '<span style="color:#9ca3af;font-size:10px">No steps assigned</span>'
        rows.append(f"""
        <div class="swimlane-row">
          <div class="swimlane-role-cell">
            <div class="swimlane-dot" style="background:{color}"></div>
            <span class="swimlane-role-name">{escape(role)}</span>
          </div>
          <div class="swimlane-steps-cell">{steps_html}</div>
        </div>""")
    return f"""
<div class="card" id="swimlanes">
  <div class="card-header"><span class="icon">🏊</span><h2>Swim Lane — Roles &amp; Responsibilities</h2></div>
  <div class="card-body" style="padding:0">
    <div class="swimlane-table">
      <div class="swimlane-header-row">
        <div class="swimlane-role-cell" style="font-weight:700;font-size:10px;text-transform:uppercase;letter-spacing:.08em;color:#6b7280">Role / Actor</div>
        <div class="swimlane-steps-cell" style="font-weight:700;font-size:10px;text-transform:uppercase;letter-spacing:.08em;color:#6b7280">Process Steps Owned</div>
      </div>
      {"".join(rows)}
    </div>
  </div>
</div>"""

# ─────────────────────────────────────────────────────────────────────────────
# L4 TABLE
# ─────────────────────────────────────────────────────────────────────────────
def build_l4_table(steps):
    if not steps:
        return '<p style="padding:20px;color:#6b7280">Steps not available.</p>'
    rows = []; prev_phase = None
    for s in steps:
        sn    = s.get("step", "1")
        phase = sn.split(".")[0]
        if phase != prev_phase:
            phase_lbl = f'<div style="font-size:9px;font-weight:700;color:#0369a1;text-transform:uppercase;letter-spacing:.07em;margin-bottom:2px">Phase {phase}</div>'
            step_cell = f"{phase_lbl}{escape(sn)}"
            row_cls   = ' class="phase-break"'; prev_phase = phase
        else:
            step_cell = escape(sn); row_cls = ""
        dp = s.get("decision_point", "N"); ex = s.get("exception", "N")
        pp = s.get("pain_point", "")
        rows.append(f"""<tr{row_cls}>
  <td class="td-step">{step_cell}</td>
  <td class="td-name">{escape(s.get("name",""))}</td>
  <td class="td-role">{escape(s.get("role",""))}</td>
  <td>{stag(s.get("system","—"))}</td>
  <td style="font-size:10px;color:#6b7280">{escape(s.get("input",""))}</td>
  <td style="font-size:10px">{escape(s.get("output",""))}</td>
  <td style="font-family:var(--mono);font-size:9px;color:#374151">{escape(s.get("kpi",""))}</td>
  <td style="text-align:center">{flag(dp)}</td>
  <td style="text-align:center">{flag(ex)}</td>
  {"<td style='font-size:9px;color:#dc2626;font-style:italic'>"+escape(pp)+"</td>" if pp else "<td></td>"}
</tr>""")
    return f"""<div class="table-wrap"><table>
<thead><tr>
  <th>Step</th><th>Step Name</th><th>Role</th><th>System</th>
  <th>Input</th><th>Output</th><th>KPI</th><th>Dec?</th><th>Exc?</th><th>Pain Point</th>
</tr></thead>
<tbody>{"".join(rows)}</tbody></table></div>"""

# ─────────────────────────────────────────────────────────────────────────────
# PROCESS ATTRIBUTES META CARD
# ─────────────────────────────────────────────────────────────────────────────
def build_meta(proc, data, steps):
    pid    = proc["id"]
    roles  = list(dict.fromkeys(s.get("role","") for s in steps if s.get("role","")))
    syss   = list(dict.fromkeys(s.get("system","") for s in steps if s.get("system","")))
    risks  = data.get("risks", [])
    kpis   = data.get("kpis", [])
    phases = list(dict.fromkeys(s.get("step","1").split(".")[0] for s in steps))
    n_gates = sum(1 for s in steps if s.get("decision_point","N") == "Y")

    swim_html = "".join(
        f'<div class="swim"><div class="dot" style="background:{SWIM_COLOURS[i%len(SWIM_COLOURS)]}"></div>{escape(r)}</div>'
        for i, r in enumerate(roles))
    sys_html  = "".join(stag(s) for s in syss)
    risk_html = "".join(f'<div class="risk-card"><strong>Risk:</strong> {escape(r)}</div>' for r in risks[:5])
    kpi_html  = "".join(f'<div class="krow"><span class="kname">•</span><span class="kval" style="text-align:left;font-family:var(--font)">{escape(k)}</span></div>' for k in kpis[:6])
    first = steps[0] if steps else {}; last = steps[-1] if steps else {}

    return f"""<div class="card" id="meta">
  <div class="card-header"><span class="icon">📋</span><h2>Process Attributes</h2></div>
  <div class="card-body"><div class="meta-grid">
    <div>
      <div class="meta-section"><h3>Identification</h3>
        <div class="krow"><span class="kname">Process ID</span><span class="kval">{pid}</span></div>
        <div class="krow"><span class="kname">L1 Domain</span><span class="kval">{escape(proc["l1_domain"])}</span></div>
        <div class="krow"><span class="kname">L2 Process Group</span><span class="kval">{escape(proc["l2_process"])}</span></div>
        <div class="krow"><span class="kname">L3 Process Name</span><span class="kval">{escape(proc["l3_name"])}</span></div>
        <div class="krow"><span class="kname">L4 Steps</span><span class="kval">{len(steps)} across {len(phases)} phases</span></div>
        <div class="krow"><span class="kname">Decision Gates</span><span class="kval">{n_gates}</span></div>
        <div class="krow"><span class="kname">Trigger</span><span class="kval" style="text-align:left;font-family:var(--font);font-size:10px">{escape(data.get("trigger",""))}</span></div>
        <div class="krow"><span class="kname">Primary Input</span><span class="kval">{escape(first.get("input","—"))}</span></div>
        <div class="krow"><span class="kname">Primary Output</span><span class="kval">{escape(last.get("output","—"))}</span></div>
      </div>
      <div class="meta-section"><h3>Swim Lanes (Roles)</h3>{swim_html}</div>
      <div class="meta-section"><h3>Systems &amp; Tools</h3>{sys_html}</div>
    </div>
    <div>
      <div class="meta-section"><h3>Key Performance Indicators</h3>{kpi_html}</div>
      <div class="meta-section"><h3>RCCL-Specific Risks &amp; Pain Points</h3>{risk_html}</div>
      <div class="meta-section"><h3>Process Description</h3>
        <p style="font-size:12px;color:#374151;line-height:1.7">{escape(data.get("description",""))}</p>
        <p style="font-size:11px;color:#6b7280;margin-top:8px"><strong>Outcome:</strong> {escape(data.get("outcome",""))}</p>
      </div>
    </div>
  </div></div></div>"""

# ─────────────────────────────────────────────────────────────────────────────
# PREV / NEXT NAV
# ─────────────────────────────────────────────────────────────────────────────
def build_pn_nav(proc):
    ids   = [p["id"] for p in ALL_PROCS]
    idx   = ids.index(proc["id"])
    prev_p = ALL_PROCS[idx - 1] if idx > 0 else None
    next_p = ALL_PROCS[idx + 1] if idx < len(ALL_PROCS) - 1 else None
    prev_html = (
        f'<a class="pn-btn" href="../{prev_p["id"].lower()}/">'
        f'<span class="arrow">←</span>'
        f'<span><span class="sub">Previous</span>{prev_p["id"]} · {escape(prev_p["l3_name"])}</span></a>'
    ) if prev_p else '<span></span>'
    next_html = (
        f'<a class="pn-btn" href="../{next_p["id"].lower()}/">'
        f'<span><span class="sub">Next</span>{next_p["id"]} · {escape(next_p["l3_name"])}</span>'
        f'<span class="arrow">→</span></a>'
    ) if next_p else '<span></span>'
    return f'<div class="pn-nav">{prev_html}{next_html}</div>'

# ─────────────────────────────────────────────────────────────────────────────
# PAGE HTML BUILDER
# ─────────────────────────────────────────────────────────────────────────────
def generate_page_html(proc, data):
    pid      = proc["id"]
    steps    = data.get("l4_steps", [])
    n_steps  = len(steps)
    phases   = list(dict.fromkeys(s.get("step","1").split(".")[0] for s in steps))
    n_gates  = sum(1 for s in steps if s.get("decision_point","N") == "Y")
    now_str  = datetime.now().strftime("%Y-%m-%d %H:%M")
    l1s      = proc["l1_slug"]; l2s = proc["l2_slug"]

    sidebar_html  = build_sidebar(pid, depth=3)
    swimlane_html = build_swimlane_section(data.get("swim_lanes",[]), steps)
    l4_html       = build_l4_table(steps)
    meta_html     = build_meta(proc, data, steps)
    pn_html       = build_pn_nav(proc)

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>{pid} · {escape(proc["l3_name"])} · Cruise Process Wiki</title>
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=IBM+Plex+Mono:wght@400;500&display=swap" rel="stylesheet">
  <link rel="stylesheet" href="../../../assets/css/wiki.css">
</head>
<body>
<header class="topbar">
  <a href="../../../" class="topbar-logo">Cruise <span>Process Wiki</span></a>
  <span class="topbar-badge">v1.0</span>
  <div class="topbar-right">
    <div class="topbar-search-wrap">
      <input type="text" id="searchBox" class="search-box" placeholder="Search processes… (/)" autocomplete="off">
      <div class="search-results" id="searchResults"></div>
    </div>
    <a href="../../../">Home</a>
    <a href="../../../{l1s}/">{escape(proc["l1_domain"])}</a>
    <a href="https://github.com/{GH_REPO}" class="gh-btn" target="_blank">⭐ GitHub</a>
  </div>
</header>

<nav class="sidebar" id="sidebar">
{sidebar_html}
</nav>
<button id="sidebarToggle" title="Toggle sidebar">◀</button>

<main class="main">
  <div class="breadcrumb">
    <a href="../../../">Home</a><span class="sep">›</span>
    <a href="../../../{l1s}/">{escape(proc["l1_domain"])}</a><span class="sep">›</span>
    <a href="../../../{l1s}/{l2s}/">{escape(proc["l2_process"])}</a><span class="sep">›</span>
    <span class="current">{pid} · {escape(proc["l3_name"])}</span>
  </div>

  <div class="page-header">
    <div class="page-header-left">
      <h1>{escape(proc["l3_name"])}</h1>
      <p>{escape(proc["l2_process"])} &nbsp;·&nbsp; {n_steps} L4 steps &nbsp;·&nbsp; {len(phases)} phases &nbsp;·&nbsp; {n_gates} decision gates &nbsp;·&nbsp; Updated {now_str}</p>
    </div>
    <div class="page-header-meta">
      <span class="pid-badge">{pid}</span>
      <span class="status-badge complete">✅ COMPLETE</span>
      <span style="font-size:9px;color:rgba(255,255,255,.5);margin-top:4px">Ref: Royal Caribbean International</span>
    </div>
  </div>

  <!-- DIAGRAM -->
  <div class="card" id="diagram">
    <div class="card-header"><span class="icon">📊</span><h2>Process Flow Diagram (BPMN)</h2></div>
    <div class="card-body">
      <div class="diagram-wrap" id="diag-{pid.lower()}">
        <img src="../../../assets/img/{pid.lower()}.png"
             alt="{pid} BPMN diagram"
             style="max-width:100%;border-radius:6px;box-shadow:0 2px 12px rgba(0,0,0,.08);cursor:zoom-in"
             onload="this.style.display='block';var ph=document.getElementById('diag-ph-{pid.lower()}');if(ph)ph.style.display='none';"
             onerror="this.style.display='none';var ph=document.getElementById('diag-ph-{pid.lower()}');if(ph)ph.style.display='flex';">
        <div id="diag-ph-{pid.lower()}" class="diagram-placeholder" style="display:none">
          <span class="ico">🗂</span>
          <span>Diagram image not yet generated.</span>
        </div>
      </div>
      <p class="mermaid-hint">Click diagram to open fullscreen &nbsp;|&nbsp; Scroll to zoom &nbsp;|&nbsp; Drag to pan</p>
    </div>
  </div>

  {swimlane_html}

  <!-- L4 TABLE -->
  <div class="card" id="l4steps">
    <div class="card-header"><span class="icon">📋</span><h2>L4 Process Steps</h2></div>
    <div class="card-body" style="padding:0">
      {l4_html}
    </div>
  </div>

  {meta_html}
  {pn_html}
</main>

<script src="../../../assets/js/wiki.js"></script>
</body>
</html>"""

# ─────────────────────────────────────────────────────────────────────────────
# GITHUB PUSH
# ─────────────────────────────────────────────────────────────────────────────
def gh_push(path, content, message):
    url     = f"{GH_API}/repos/{GH_REPO}/contents/{path}"
    headers = {"Authorization": f"token {GH_TOKEN}", "Accept": "application/vnd.github.v3+json"}
    sha = None
    try:
        r = requests.get(url, headers=headers, timeout=15)
        if r.status_code == 200: sha = r.json()["sha"]
    except: pass
    if isinstance(content, bytes):
        encoded = base64.b64encode(content).decode("ascii")
    else:
        encoded = base64.b64encode(content.encode("utf-8")).decode("ascii")
    body = {"message": message, "content": encoded}
    if sha: body["sha"] = sha
    for attempt in range(3):
        try:
            r = requests.put(url, headers=headers, json=body, timeout=60)
            if r.status_code in (200, 201): return True
            log(f"Push {r.status_code}: {r.text[:100]}", "warn")
        except Exception as ex:
            log(f"Push error: {ex}", "warn")
        time.sleep(2)
    return False

# ─────────────────────────────────────────────────────────────────────────────
# VERIFY WIKI PAGE (after 90s propagation wait)
# ─────────────────────────────────────────────────────────────────────────────
def verify_wiki_page(pid, proc):
    url = (f"https://ghatk047.github.io/Cruise-Process-Wiki/"
           f"{proc['l1_slug']}/{proc['l2_slug']}/{pid.lower()}/")
    try:
        r = requests.get(url, timeout=20)
        if r.status_code != 200:
            log(f"Wiki page {r.status_code} — not writing link", "warn"); return None
        if pid not in r.text:
            log(f"PID {pid} not found in page content", "warn"); return None
        if "assets/img" not in r.text:
            log(f"assets/img not found — page may be stale", "warn"); return None
        return url
    except Exception as ex:
        log(f"Wiki verify failed: {ex}", "warn")
    return None

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL — create catalog or update
# ─────────────────────────────────────────────────────────────────────────────
def init_excel():
    """Create fresh Excel catalog with all processes pre-loaded as Queued."""
    EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    # ── INDEX SHEET ──────────────────────────────────────────────────────────
    ws = wb.active; ws.title = "Index"
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill(fill_type="solid", fgColor="003366")
    headers  = ["Process ID","L1 Domain","L2 Process Group","L3 Process Name",
                "Status","Wiki Link","Notes","Completed Date","Steps #"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    col_widths = [12, 32, 32, 40, 10, 55, 20, 14, 8]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 20

    row = 2
    for proc in ALL_PROCS:
        ws.cell(row=row, column=1, value=proc["id"])
        ws.cell(row=row, column=2, value=proc["l1_domain"])
        ws.cell(row=row, column=3, value=proc["l2_process"])
        ws.cell(row=row, column=4, value=proc["l3_name"])
        ws.cell(row=row, column=5, value="⏳ Queued")
        row += 1

    # ── SUMMARY SHEET ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 36
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 10
    ws2["A1"] = "L1 Domain"; ws2["B1"] = "Total"; ws2["C1"] = "Complete"
    for c in ["A1","B1","C1"]:
        ws2[c].font = hdr_font; ws2[c].fill = hdr_fill
        ws2[c].alignment = Alignment(horizontal="center")

    seen = {}
    for proc in ALL_PROCS:
        d = proc["l1_domain"]
        seen[d] = seen.get(d, 0) + 1

    row2 = 2
    for dom, total in seen.items():
        ws2.cell(row=row2, column=1, value=dom)
        ws2.cell(row=row2, column=2, value=total)
        ws2.cell(row=row2, column=3, value=0)
        row2 += 1

    # ── MASTER CATALOG SHEET ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("Master Catalog")
    mc_headers = ["Process ID","L1 Domain","L2 Process","L3 Name",
                  "L4 Step #","L4 Step Name","Role / Swim Lane","System",
                  "Input","Output","KPI","Pain Point / Risk",
                  "Decision Point","Exception"]
    for col, h in enumerate(mc_headers, 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")
    for i in range(1, 15):
        ws3.column_dimensions[get_column_letter(i)].width = 18

    wb.save(EXCEL_PATH)
    log(f"Excel catalog created: {EXCEL_PATH}", "ok")
    log(f"  {len(ALL_PROCS)} processes pre-loaded as Queued")

def update_excel(pid, proc, steps):
    """Update Index row + write per-process tab + append to Master Catalog."""
    try:
        wb = load_workbook(EXCEL_PATH)
    except:
        init_excel(); wb = load_workbook(EXCEL_PATH)

    now_str = datetime.now().strftime("%Y-%m-%d")

    # ── Update Index row ──────────────────────────────────────────────────────
    ws = wb["Index"]
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value == pid:
            row[4].value = "✅ Complete"
            row[7].value = now_str
            row[8].value = len(steps)
            break

    # ── Per-process worksheet ─────────────────────────────────────────────────
    tab_name = pid[:31]
    if tab_name in wb.sheetnames:
        del wb[tab_name]
    ws2 = wb.create_sheet(tab_name)
    hdr_font = Font(bold=True, color="FFFFFF", size=9)
    hdr_fill = PatternFill(fill_type="solid", fgColor="003366")
    mc_headers = ["Process ID","L1 Domain","L2 Process","L3 Name",
                  "L4 Step #","L4 Step Name","Role / Swim Lane","System",
                  "Input","Output","KPI","Pain Point / Risk",
                  "Decision Point","Exception"]
    for col, h in enumerate(mc_headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = hdr_font; c.fill = hdr_fill
    for s in steps:
        ws2.append([
            pid, proc["l1_domain"], proc["l2_process"], proc["l3_name"],
            s.get("step",""), s.get("name",""), s.get("role",""), s.get("system",""),
            s.get("input",""), s.get("output",""), s.get("kpi",""), s.get("pain_point",""),
            s.get("decision_point","N"), s.get("exception","N"),
        ])
    for i in range(1, 15):
        ws2.column_dimensions[get_column_letter(i)].width = 18
    log(f"Excel worksheet created: tab '{tab_name}'", "ok")

    # ── Master Catalog append ─────────────────────────────────────────────────
    ws3 = wb["Master Catalog"]
    for s in steps:
        ws3.append([
            pid, proc["l1_domain"], proc["l2_process"], proc["l3_name"],
            s.get("step",""), s.get("name",""), s.get("role",""), s.get("system",""),
            s.get("input",""), s.get("output",""), s.get("kpi",""), s.get("pain_point",""),
            s.get("decision_point","N"), s.get("exception","N"),
        ])
    log(f"Master Catalog updated: +{len(steps)} rows", "ok")

    wb.save(EXCEL_PATH)
    log(f"Excel saved ✅")

def write_wiki_link_to_excel(pid, wiki_url):
    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb["Index"]
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[0].value == pid:
                row[5].value = wiki_url
                break
        wb.save(EXCEL_PATH)
        log(f"Wiki link written to Excel: {wiki_url}", "ok")
    except Exception as ex:
        log(f"Could not write wiki link: {ex}", "warn")

# ─────────────────────────────────────────────────────────────────────────────
# HOME PAGE UPDATE
# ─────────────────────────────────────────────────────────────────────────────
def update_home_page():
    """Push updated home page with accurate domain completion counts."""
    total_done  = sum(1 for p in ALL_PROCS if p["status"] == "Complete")
    total_procs = len(ALL_PROCS)

    # Per-domain counts
    domain_counts = {}
    for proc in ALL_PROCS:
        d = proc["l1_domain"]
        if d not in domain_counts:
            domain_counts[d] = {"total": 0, "done": 0}
        domain_counts[d]["total"] += 1
        if proc["status"] == "Complete":
            domain_counts[d]["done"] += 1

    ICONS = {
        "Guest Services & Embarkation": "⚓",
        "Stateroom & Housekeeping": "🛏️",
        "Food & Beverage Operations": "🍽️",
        "Entertainment & Activities": "🎭",
        "Shore Excursions & Destinations": "🌴",
        "Marine & Technical Operations": "⚙️",
        "Environmental & Sustainability": "🌊",
        "Crew Management & HR": "👥",
        "Revenue & Commercial": "💰",
        "Finance & Procurement": "💼",
        "Technology & Cybersecurity": "💻",
    }

    domain_cards = ""
    for l1_name, l1_code, l1_slug, groups in TAXONOMY:
        cnt   = domain_counts.get(l1_name, {"total": 0, "done": 0})
        icon  = ICONS.get(l1_name, "🚢")
        grp_count = len(groups)
        domain_cards += f"""    <div class="domain-card">
      <div class="domain-card-hdr"><h3>{icon} {escape(l1_name)}</h3><p>Cruise · {grp_count} process group{"s" if grp_count != 1 else ""}</p></div>
      <div class="domain-card-body">
        <div class="domain-card-stat">{cnt["done"]} of {cnt["total"]} subprocesses complete</div>
        <a class="domain-card-link" href="{l1_slug}/">Explore domain →</a>
      </div>
    </div>\n"""

    sidebar_html = build_sidebar("", depth=0)

    index_html = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>Cruise Process Wiki — Royal Caribbean Reference</title>
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=IBM+Plex+Mono:wght@400;500&display=swap" rel="stylesheet">
  <link rel="stylesheet" href="./assets/css/wiki.css">
</head>
<body>
<header class="topbar">
  <a href="./" class="topbar-logo">Cruise <span>Process Wiki</span></a>
  <span class="topbar-badge">v1.0</span>
  <div class="topbar-right">
    <div class="topbar-search-wrap">
      <input type="text" id="searchBox" class="search-box" placeholder="Search processes… (/)" autocomplete="off">
      <div class="search-results" id="searchResults"></div>
    </div>
    <a href="./">Home</a>
    <a href="https://github.com/{GH_REPO}" class="gh-btn" target="_blank">⭐ GitHub</a>
  </div>
</header>
<nav class="sidebar" id="sidebar">
{sidebar_html}
</nav>
<button id="sidebarToggle" title="Toggle sidebar">◀</button>
<main class="main">
  <div class="hero">
    <div>
      <div class="hero-label">Process Catalog</div>
      <h1 class="hero-title">Cruise Process Wiki</h1>
      <p class="hero-desc">End-to-end L1 → L2 → L3 → L4 process documentation for cruise & integrated resort operations. Each subprocess includes BPMN process flow diagrams, L4 step tables, swim lane documentation, system landscape (Royal Caribbean International reference), KPIs, and cruise-specific risk analysis.</p>
      <div class="hero-stats">
        <span>✅ {total_done} subprocesses complete</span>
        <span>📋 {total_procs} total processes</span>
        <span>🚢 Ref: Royal Caribbean International</span>
      </div>
    </div>
  </div>
  <div style="font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.1em;color:#6b7280;margin-bottom:12px">L1 DOMAINS</div>
  <div class="domain-grid">
{domain_cards}  </div>
</main>
<script src="./assets/js/wiki.js"></script>
</body>
</html>"""

    ok = gh_push("index.html", index_html,
                 f"Update home page: {total_done}/{total_procs} complete")
    if ok:
        log(f"Home page updated: {total_done}/{total_procs} complete", "ok")
    else:
        log("Home page push failed", "warn")

# ─────────────────────────────────────────────────────────────────────────────
# L1 / L2 INDEX PAGES
# NOTE: These are pushed ONLY at the end of a pilot/batch run, not after
# every process. This avoids 20+ GitHub API calls per process (performance fix).
# The home page counter IS updated after every process.
# ─────────────────────────────────────────────────────────────────────────────
def push_all_index_pages():
    """Generate and push all L1 and L2 index pages. Called ONCE after batch."""
    pushed = 0
    for l1_name, l1_code, l1_slug, groups in TAXONOMY:
        groups_data = []
        for l2_name, l2_code, l2_slug, proc_names in groups:
            procs = [PROC_MAP[f"{l2_code}-{i+1:02d}"] for i in range(len(proc_names))]
            groups_data.append((l2_name, l2_code, l2_slug, procs))

            # L2 index
            done  = sum(1 for p in procs if p.get("status") == "Complete")
            total = len(procs)
            sidebar_html = build_sidebar("", depth=2)
            cards = ""
            for p in procs:
                pid_  = p["id"]
                status = p.get("status","Queued")
                dot   = "✅" if status == "Complete" else "⏳"
                steps = len(p.get("l4_steps",[]))
                step_txt = f"{steps} L4 steps" if steps else "pending"
                cards += f"""    <div class="domain-card">
      <div class="domain-card-hdr"><h3>{escape(p["l3_name"])}</h3><p>{pid_}</p></div>
      <div class="domain-card-body">
        <div class="domain-card-stat">{dot} {status} · {step_txt}</div>
        <a class="domain-card-link" href="{pid_.lower()}/">View process →</a>
      </div>
    </div>\n"""
            l2_html = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
  <title>{escape(l2_name)} · Cruise Process Wiki</title>
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=IBM+Plex+Mono:wght@400;500&display=swap" rel="stylesheet">
  <link rel="stylesheet" href="../../assets/css/wiki.css">
</head>
<body>
<header class="topbar">
  <a href="../../" class="topbar-logo">Cruise <span>Process Wiki</span></a>
  <span class="topbar-badge">v1.0</span>
  <div class="topbar-right">
    <a href="../../">Home</a>
    <a href="../../{l1_slug}/">{escape(l1_name)}</a>
    <a href="https://github.com/{GH_REPO}" class="gh-btn" target="_blank">⭐ GitHub</a>
  </div>
</header>
<nav class="sidebar" id="sidebar">{sidebar_html}</nav>
<button id="sidebarToggle" title="Toggle sidebar">◀</button>
<main class="main">
  <div class="breadcrumb">
    <a href="../../">Home</a><span class="sep">›</span>
    <a href="../../{l1_slug}/">{escape(l1_name)}</a><span class="sep">›</span>
    <span class="current">{escape(l2_name)}</span>
  </div>
  <div class="page-header">
    <div class="page-header-left">
      <h1>{escape(l2_name)}</h1>
      <p>{escape(l1_name)} · {total} processes · {done} complete · Reference: Royal Caribbean International</p>
    </div>
    <div class="page-header-meta">
      <span class="pid-badge">{l2_code}</span>
      <span class="status-badge {'complete' if done==total and total>0 else 'queued'}">{done}/{total} complete</span>
    </div>
  </div>
  <div class="domain-grid">{cards}</div>
</main>
<script src="../../assets/js/wiki.js"></script>
</body></html>"""
            if gh_push(f"{l1_slug}/{l2_slug}/index.html", l2_html,
                       f"Update {l2_code} index page"):
                pushed += 1

        # L1 index
        done_all  = sum(sum(1 for p in g[3] if p.get("status")=="Complete") for g in groups_data)
        total_all = sum(len(g[3]) for g in groups_data)
        sidebar_html = build_sidebar("", depth=1)
        cards = ""
        for l2_name, l2_code, l2_slug, procs in groups_data:
            done  = sum(1 for p in procs if p.get("status") == "Complete")
            total = len(procs)
            cards += f"""    <div class="domain-card">
      <div class="domain-card-hdr"><h3>{escape(l2_name)}</h3><p>{total} processes</p></div>
      <div class="domain-card-body">
        <div class="domain-card-stat">{done} of {total} subprocesses complete</div>
        <a class="domain-card-link" href="{l2_slug}/">Explore group →</a>
      </div>
    </div>\n"""
        l1_html = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
  <title>{escape(l1_name)} · Cruise Process Wiki</title>
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=IBM+Plex+Mono:wght@400;500&display=swap" rel="stylesheet">
  <link rel="stylesheet" href="../assets/css/wiki.css">
</head>
<body>
<header class="topbar">
  <a href="../" class="topbar-logo">Cruise <span>Process Wiki</span></a>
  <span class="topbar-badge">v1.0</span>
  <div class="topbar-right">
    <a href="../">Home</a>
    <a href="https://github.com/{GH_REPO}" class="gh-btn" target="_blank">⭐ GitHub</a>
  </div>
</header>
<nav class="sidebar" id="sidebar">{sidebar_html}</nav>
<button id="sidebarToggle" title="Toggle sidebar">◀</button>
<main class="main">
  <div class="breadcrumb">
    <a href="../">Home</a><span class="sep">›</span>
    <span class="current">{escape(l1_name)}</span>
  </div>
  <div class="page-header">
    <div class="page-header-left">
      <h1>{escape(l1_name)}</h1>
      <p>Cruise Process Wiki · {len(groups_data)} process groups · {done_all} of {total_all} complete · Reference: Royal Caribbean International</p>
    </div>
    <div class="page-header-meta">
      <span class="pid-badge">{l1_code}</span>
      <span class="status-badge {'complete' if done_all==total_all and total_all>0 else 'queued'}">{done_all}/{total_all} complete</span>
    </div>
  </div>
  <div style="font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.1em;color:#6b7280;margin-bottom:12px">Process Groups</div>
  <div class="domain-grid">{cards}</div>
</main>
<script src="../assets/js/wiki.js"></script>
</body></html>"""
        if gh_push(f"{l1_slug}/index.html", l1_html,
                   f"Update {l1_code} ({l1_name}) index page"):
            pushed += 1

    log(f"Index pages pushed: {pushed} files", "ok")

# ─────────────────────────────────────────────────────────────────────────────
# MAIN RUN LOOP
# ─────────────────────────────────────────────────────────────────────────────
def run(args):
    model      = args.model
    start_id   = args.start
    only_id    = args.only
    max_run    = args.max or 9999
    pilot_only = args.pilot

    log("=" * 60, "head")
    log("Cruise Process Wiki Generator", "head")
    log(f"   Model : {model}")
    log(f"   Repo  : https://github.com/{GH_REPO}")
    log(f"   Excel : {EXCEL_PATH}")
    log("=" * 60, "head")

    load_progress()

    # Ensure Excel exists
    if not EXCEL_PATH.exists():
        log("Excel catalog not found — creating fresh catalog", "warn")
        init_excel()

    # Build queue
    if only_id:
        queue = [p for p in ALL_PROCS if p["id"] == only_id]
    elif pilot_only:
        queue = [p for p in ALL_PROCS if p["status"] != "Complete"][:2]
    else:
        queue = [p for p in ALL_PROCS if p["status"] != "Complete"]
        if start_id:
            ids = [p["id"] for p in queue]
            if start_id in ids:
                queue = queue[ids.index(start_id):]
            else:
                log(f"--start ID '{start_id}' not found in queue", "warn")

    queue = queue[:max_run]
    log(f"   Queue : {len(queue)} processes to generate")
    log("-" * 60)

    done = 0; failed = 0
    t_total = time.time()

    for i, proc in enumerate(queue, 1):
        pid = proc["id"]
        log(f"[{i}/{len(queue)}] {pid} — {proc['l3_name']}", "step")

        # 1. Call Qwen
        prompt = build_prompt(proc)
        data   = call_qwen(prompt, model)
        if not data:
            log(f"SKIPPED {pid} — all retries failed", "err")
            failed += 1; continue

        steps = data.get("l4_steps", [])
        log(f"Generated {len(steps)} steps, {len(data.get('swim_lanes',[]))} swim lanes", "ok")

        # 2. Sanitise Mermaid & render PNG via mmdc
        mmd_text = sanitise_mmd(data.get("mermaid", ""))
        png_ok   = generate_png(pid, mmd_text)

        # 3. Build and push HTML
        proc["status"] = "Complete"
        html = generate_page_html(proc, data)
        path = f"{proc['l1_slug']}/{proc['l2_slug']}/{pid.lower()}/index.html"
        log(f"Pushing {path}...", "push")
        ok = gh_push(path, html, f"Add {pid}: {proc['l3_name']}")
        if ok:
            log(f"Pushed → https://ghatk047.github.io/Cruise-Process-Wiki/{path}", "ok")
        else:
            log(f"Push failed for {pid}", "err")

        # 4. Update Excel (with wiki link written after verification)
        update_excel(pid, proc, steps)
        proc["l4_steps"] = steps
        save_progress()

        # 5. Update home page counter (lightweight — just index.html)
        update_home_page()

        # 6. Verify wiki page after GitHub Pages propagation (90s)
        log("Waiting 90s for GitHub Pages to propagate...", "info")
        time.sleep(90)
        wiki_url = verify_wiki_page(pid, proc)
        if wiki_url:
            log(f"Wiki link verified: {wiki_url}", "ok")
            write_wiki_link_to_excel(pid, wiki_url)
        else:
            log(f"Wiki link NOT written (page not live yet)", "warn")

        done += 1
        elapsed   = time.time() - t_total
        avg       = elapsed / done
        remaining = avg * (len(queue) - i)
        log(f"Progress: {done}/{len(queue)} done | ~{remaining/60:.0f} min remaining")
        log("-" * 50)

        if i < len(queue):
            time.sleep(1)

    # Push all L1+L2 index pages ONCE at the end (not after every process)
    log("Pushing all L1 + L2 index pages...", "push")
    push_all_index_pages()

    # Final .deploy marker for clean GitHub Actions deployment
    gh_push(".deploy",
            f"cruise-process-wiki\ndeployed: {datetime.now().isoformat()}\nprocesses-done: {done}",
            f"Deploy marker: {done} processes complete")

    log("=" * 60)
    log(f"DONE — {done} generated, {failed} failed", "head")
    log(f"Total time: {(time.time()-t_total)/60:.1f} minutes")
    log(f"Live wiki: https://ghatk047.github.io/Cruise-Process-Wiki/")
    log("=" * 60)

# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    p = argparse.ArgumentParser(description="Cruise Process Wiki Generator — Royal Caribbean reference")
    p.add_argument("--model",     default="qwen2.5-coder:14b", help="Ollama model name")
    p.add_argument("--start",     default=None,  help="Resume from this process ID")
    p.add_argument("--only",      default=None,  help="Generate only this one process ID")
    p.add_argument("--max",       default=None,  type=int, help="Max processes to run this session")
    p.add_argument("--pilot",     action="store_true", help="Run first 2 queued processes only")
    p.add_argument("--reset",     default=None,  help="Reset comma-separated IDs to Queued")
    p.add_argument("--reset-all", action="store_true", help="Reset ALL processes to Queued")
    args = p.parse_args()

    if args.reset_all:
        load_progress()
        for proc in ALL_PROCS:
            proc["status"] = "Queued"; proc["l4_steps"] = []
        save_progress()
        print(f"[{ts()}] ✅ Reset ALL {len(ALL_PROCS)} processes to Queued")
        sys.exit(0)

    if args.reset:
        load_progress()
        ids_to_reset = [x.strip() for x in args.reset.split(",")]
        n = 0
        for proc in ALL_PROCS:
            if proc["id"] in ids_to_reset:
                proc["status"] = "Queued"; proc["l4_steps"] = []; n += 1
                print(f"[{ts()}]    Reset {proc['id']} → Queued")
        save_progress()
        print(f"[{ts()}] ✅ Reset {n} processes to Queued")
        sys.exit(0)

    run(args)
